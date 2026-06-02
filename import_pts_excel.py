#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Import: สถานที่ล้างแอร์ + พัดลม.xlsx  →  airclean_db
Run: python import_pts_excel.py
"""

import os
import re
import sys
import calendar
from datetime import date

import openpyxl
import psycopg2

# ════════════════════════════ CONFIG ════════════════════════════
# All connection settings + secrets come from environment (see .env.example).
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

EXCEL_FILE = os.environ.get('EXCEL_PATH', '')

DB = dict(
    host     = os.environ.get('DB_HOST', 'localhost'),
    port     = int(os.environ.get('DB_PORT', '5432')),
    dbname   = os.environ.get('DB_NAME', 'airclean_db'),
    user     = os.environ.get('DB_USER', 'postgres'),
    password = os.environ.get('DB_PASS'),
)

if not DB['password']:
    sys.exit("ERROR: DB_PASS not set. Configure env vars (see .env.example).")
if not EXCEL_FILE:
    sys.exit("ERROR: EXCEL_PATH not set. Configure env vars (see .env.example).")

PM_INTERVAL = 2   # default เดือน

# work columns (0-based index from col A)
WORK_COLS = [
    (6, 'major'),   # ล้างใหญ่ ครั้งที่ 1
    (7, 'minor'),   # ล้างย่อย ครั้งที่ 1
    (8, 'minor'),   # ล้างย่อย ครั้งที่ 2
    (9, 'major'),   # ล้างใหญ่ ครั้งที่ 2
    (10, 'minor'),  # ล้างย่อย ครั้งที่ 3
    (11, 'minor'),  # ล้างย่อย ครั้งที่ 4
]

SKIP_VALS  = {'แอร์เสีย', 'ไม่มีฟิลเตอร์', 'ถอดออก', 'รื้อถอน', '-', ''}
BROKEN_VALS = {'แอร์เสีย', 'ถอดออก', 'รื้อถอน'}

# ════════════════════════════ HELPERS ════════════════════════════

def add_months(d, n):
    m = d.month + n
    y = d.year + (m - 1) // 12
    m = (m - 1) % 12 + 1
    day = min(d.day, calendar.monthrange(y, m)[1])
    return date(y, m, day)


def parse_ac_type(s):
    """Map raw type string → schema type: FCU / Split / AHU / VRF / Fan"""
    s = (s or '').upper()
    if 'AHU' in s:   return 'AHU'
    if 'VRF' in s:   return 'VRF'
    if 'SPLIT' in s: return 'Split'
    if 'FCU' in s:   return 'FCU'
    return 'FCU'


def parse_btu(s):
    m = re.search(r'([\d,]+)\s*BTU', str(s or ''), re.IGNORECASE)
    return str(int(m.group(1).replace(',', ''))) if m else None


def parse_date(val):
    """Return CE date or None.  Handles: BE string 'YYYY-MM-DD', datetime, date obj."""
    if val is None:
        return None

    # Python date/datetime
    if isinstance(val, date):
        d = val
        return d.replace(year=d.year - 543) if d.year > 2300 else d
    if hasattr(val, 'date'):
        d = val.date()
        return d.replace(year=d.year - 543) if d.year > 2300 else d

    s = str(val).strip()
    if not s or s.lower() in SKIP_VALS:
        return None

    # "2569-01-27"
    m = re.fullmatch(r'(\d{4})-(\d{2})-(\d{2})', s)
    if m:
        y, mo, dy = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y > 2300: y -= 543
        try:    return date(y, mo, dy)
        except: return None

    # "27/01/2569" or "27/01/2026"
    m = re.fullmatch(r'(\d{1,2})/(\d{1,2})/(\d{4})', s)
    if m:
        dy, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if y > 2300: y -= 543
        try:    return date(y, mo, dy)
        except: return None

    return None


def extract_hospital(text):
    """
    'รายการ เครื่องปรับอากาศ พญาไท ศรีราชา 1 อาคาร A'
    → 'พญาไท ศรีราชา 1'
    """
    s = str(text).strip()
    m = re.search(r'(?:เครื่องปรับอากาศ|พัดลม)\s+(.+?)\s+อาคาร\s+\w', s)
    return m.group(1).strip() if m else None


def cell_str(val):
    return str(val).strip() if val is not None else ''


# ════════════════════════════ DB HELPERS ════════════════════════

_cache: dict = {}


def _get_or_create(cur, cache_key, select_sql, select_params, insert_sql, insert_params):
    if cache_key in _cache:
        return _cache[cache_key]
    cur.execute(select_sql, select_params)
    row = cur.fetchone()
    if row:
        _cache[cache_key] = row[0]
        return row[0]
    cur.execute(insert_sql, insert_params)
    _cache[cache_key] = cur.fetchone()[0]
    return _cache[cache_key]


def get_hospital(cur, name):
    slug = re.sub(r'[^\w]', '-', name.lower())[:40] + f'-{abs(hash(name)) % 9999}'
    return _get_or_create(
        cur, f'H:{name.lower()}',
        "SELECT id FROM hospitals WHERE LOWER(name)=LOWER(%s)", (name,),
        "INSERT INTO hospitals (name, slug) VALUES (%s,%s) RETURNING id", (name, slug),
    )


def get_building(cur, h_id, name):
    return _get_or_create(
        cur, f'B:{h_id}:{name.lower()}',
        "SELECT id FROM buildings WHERE hospital_id=%s AND LOWER(name)=LOWER(%s)", (h_id, name),
        "INSERT INTO buildings (hospital_id, name, code) VALUES (%s,%s,%s) RETURNING id", (h_id, name, name[:20]),
    )


def get_floor(cur, b_id, name):
    return _get_or_create(
        cur, f'F:{b_id}:{name.lower()}',
        "SELECT id FROM floors WHERE building_id=%s AND LOWER(name)=LOWER(%s)", (b_id, name),
        "INSERT INTO floors (building_id, name) VALUES (%s,%s) RETURNING id", (b_id, name),
    )


def get_dept(cur, f_id, name):
    return _get_or_create(
        cur, f'D:{f_id}:{name.lower()}',
        "SELECT id FROM departments WHERE floor_id=%s AND LOWER(name)=LOWER(%s)", (f_id, name),
        "INSERT INTO departments (floor_id, name) VALUES (%s,%s) RETURNING id", (f_id, name),
    )


def upsert_ac(cur, dept_id, ac_code, dept_name, ac_type, btu_str, status):
    cur.execute("SELECT id, status FROM ac_units WHERE ac_code=%s", (ac_code,))
    row = cur.fetchone()
    if row:
        if status == 'broken':
            cur.execute("UPDATE ac_units SET status='broken', updated_at=NOW() WHERE id=%s", (row[0],))
        return row[0]
    cur.execute("""
        INSERT INTO ac_units
            (department_id, ac_code, name, type, capacity_btu, pm_interval_months, status)
        VALUES (%s,%s,%s,%s,%s,%s,%s) RETURNING id
    """, (dept_id, ac_code, dept_name, ac_type, btu_str, PM_INTERVAL, status))
    return cur.fetchone()[0]


# ════════════════════════════ MAIN ════════════════════════════

def main():
    print("Reading Excel:", EXCEL_FILE)
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    conn = psycopg2.connect(**DB)
    conn.autocommit = False
    cur = conn.cursor()

    stats = {'acs': 0, 'wos': 0, 'wo_errors': []}

    # ── Collect all data first ───────────────────────────────────
    # wo_groups: (hosp_id, type, date_iso) → {'hosp_id', 'type', 'date', 'items': set}
    wo_groups: dict = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        is_fan = 'พัดลม' in sheet_name
        print(f"\n[Sheet] {sheet_name}  (fan={is_fan})")

        hosp_id = None

        for row in ws.iter_rows(values_only=True):
            if all(v is None for v in row):
                continue

            col0   = row[0]
            col0s  = cell_str(col0)
            row_len = len(row)

            # ── Hospital header (col0 has hospital text, col1..col5 all null) ──
            if col0 is not None and row_len >= 6 and all(v is None for v in row[1:6]):
                h = extract_hospital(col0s)
                if h:
                    hosp_id = get_hospital(cur, h)
                    print(f"  Hospital: {h!r}  id={hosp_id}")
                continue

            # ── Skip header rows ─────────────────────────────────────────────
            if col0s in ('ลำดับ', 'ลำดับที่', 'No', 'No.') or col0s == 'อาคาร':
                continue

            # ── Data row: need valid AC code in col index 5 ─────────────────
            ac_code_raw = row[5] if row_len > 5 else None
            if not ac_code_raw:
                continue
            ac_code = cell_str(ac_code_raw)
            if not re.match(r'^AC-', ac_code):
                continue
            if hosp_id is None:
                print(f"  WARN: no hospital context for {ac_code}, skip")
                continue

            # Parse fields
            bld_raw  = cell_str(row[1]) if row_len > 1 else 'A'
            fl_raw   = cell_str(row[2]) if row_len > 2 else '1'
            dept_raw = cell_str(row[3]) if row_len > 3 else ac_code
            type_raw = cell_str(row[4]) if row_len > 4 else ''

            bld_name   = f"อาคาร {bld_raw}"   if bld_raw  else 'อาคาร A'
            floor_name = f"ชั้น {fl_raw}"      if fl_raw   else 'ชั้น 1'
            dept_name  = dept_raw              if dept_raw else ac_code
            ac_type    = 'Fan' if is_fan else parse_ac_type(type_raw)
            btu_str    = parse_btu(type_raw)

            # Broken?
            is_broken = any(
                cell_str(row[ci]) in BROKEN_VALS
                for ci, _ in WORK_COLS if ci < row_len and row[ci] is not None
            )

            # Create hierarchy + AC unit
            b_id  = get_building(cur, hosp_id, bld_name)
            f_id  = get_floor(cur, b_id, floor_name)
            d_id  = get_dept(cur, f_id, dept_name)
            ac_id = upsert_ac(cur, d_id, ac_code, dept_name, ac_type, btu_str,
                              'broken' if is_broken else 'active')
            stats['acs'] += 1

            # Collect work order dates
            for ci, wtype in WORK_COLS:
                if ci >= row_len:
                    continue
                cell_val = row[ci]
                cell_s   = cell_str(cell_val)
                if cell_s.lower() in SKIP_VALS or not cell_s:
                    continue
                d = parse_date(cell_val)
                if d is None:
                    continue

                gtype = 'fan' if is_fan else wtype
                gkey  = (hosp_id, gtype, d.isoformat())
                if gkey not in wo_groups:
                    wo_groups[gkey] = {
                        'hosp_id': hosp_id,
                        'type':    gtype,
                        'date':    d,
                        'items':   set(),
                    }
                wo_groups[gkey]['items'].add(ac_id)

    # Commit AC unit data
    conn.commit()
    print(f"\nAC units done: {stats['acs']}")

    # ── Create Work Orders ───────────────────────────────────────
    print(f"Creating {len(wo_groups)} work orders...")
    for gkey, g in wo_groups.items():
        try:
            d  = g['date']
            ds = d.strftime('%Y%m%d')

            cur.execute(
                "SELECT COUNT(*)::int FROM work_orders WHERE order_no LIKE %s",
                (f'IMP-{ds}-%',)
            )
            cnt      = cur.fetchone()[0]
            order_no = f'IMP-{ds}-{str(cnt + 1).zfill(4)}'

            cur.execute("""
                INSERT INTO work_orders
                    (order_no, hospital_id, type, status,
                     started_at, completed_at, approved_at)
                VALUES (%s,%s,%s,'approved',%s,%s,%s) RETURNING id
            """, (order_no, g['hosp_id'], g['type'], d, d, d))
            wo_id = cur.fetchone()[0]

            for ac_id in g['items']:
                # Insert item (avoid dup)
                cur.execute("""
                    INSERT INTO work_order_items (work_order_id, ac_unit_id)
                    SELECT %s, %s
                    WHERE NOT EXISTS (
                        SELECT 1 FROM work_order_items
                        WHERE work_order_id=%s AND ac_unit_id=%s
                    )
                """, (wo_id, ac_id, wo_id, ac_id))

                # Update next_pm_date
                next_d = add_months(d, PM_INTERVAL)
                cur.execute("""
                    UPDATE ac_units
                    SET next_pm_date = GREATEST(COALESCE(next_pm_date,%s::date), %s::date),
                        updated_at   = NOW()
                    WHERE id=%s
                """, (next_d, next_d, ac_id))

            conn.commit()
            stats['wos'] += 1

        except Exception as e:
            conn.rollback()
            stats['wo_errors'].append(f"{gkey}: {e}")

    cur.close()
    conn.close()

    print(f"\n{'='*50}")
    print(f"AC units processed : {stats['acs']}")
    print(f"Work orders created: {stats['wos']}")
    if stats['wo_errors']:
        print(f"WO errors ({len(stats['wo_errors'])}):")
        for err in stats['wo_errors']:
            print(f"  {err}")
    print("Done!")


if __name__ == '__main__':
    main()
