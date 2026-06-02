"""
Import AC units from Excel to airclean_db
Files:
  - สถานที่ล้างแอร์ + พัดลม.xlsx  (sheet: บันทึกวันล้างแอร์ 2569)
  - สถานที่ล้างแอร์ + พัดลม.xlsx  (sheet: บันทึกวันล้างพัดลม 2569)
"""

import openpyxl
import psycopg2
import sys
import os
from datetime import datetime

# ── Config (read from environment — never hardcode secrets) ─────────────────────
# Optionally load a local .env file if python-dotenv is installed.
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass

DB = {
    'host':     os.environ.get('DB_HOST', 'localhost'),
    'port':     int(os.environ.get('DB_PORT', '5432')),
    'dbname':   os.environ.get('DB_NAME', 'airclean_db'),
    'user':     os.environ.get('DB_USER', 'postgres'),
    'password': os.environ.get('DB_PASS'),
}
EXCEL_PATH = os.environ.get('EXCEL_PATH', '')
HOSPITAL_SLUG = os.environ.get('HOSPITAL_SLUG', 'pts1')   # โรงพยาบาลพญาไท ศรีราชา 1

if not DB['password']:
    sys.exit("ERROR: DB_PASS not set. Configure env vars (see .env.example).")
if not EXCEL_PATH:
    sys.exit("ERROR: EXCEL_PATH not set. Configure env vars (see .env.example).")

# ── AC type mapping ───────────────────────────────────────────────────────────
def map_type(raw: str) -> str:
    if not raw:
        return 'FCU'
    r = str(raw).lower()
    if 'fcu' in r or 'fan coil' in r:
        return 'FCU'
    if 'split' in r:
        return 'Split'
    if 'ahu' in r:
        return 'AHU'
    if 'vrf' in r:
        return 'VRF'
    if 'พัดลม' in r or 'fan' in r:
        return 'Fan'
    return 'Split'

def map_status(raw) -> str:
    if raw is None:
        return 'active'
    s = str(raw).strip()
    if 'เสีย' in s:
        return 'broken'
    if 'ไม่มี' in s:
        return 'active'
    if isinstance(raw, datetime):
        return 'active'
    return 'active'

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    conn = psycopg2.connect(**DB)
    cur = conn.cursor()

    # get hospital id
    cur.execute("SELECT id FROM hospitals WHERE slug = %s", (HOSPITAL_SLUG,))
    row = conn.cursor()
    cur.execute("SELECT id FROM hospitals WHERE slug = %s", (HOSPITAL_SLUG,))
    result = cur.fetchone()
    if not result:
        print(f"Hospital '{HOSPITAL_SLUG}' not found. Run seed.js first.")
        sys.exit(1)
    hospital_id = result[0]
    print(f"Hospital id: {hospital_id}")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    total_ac = 0
    total_fan = 0

    # ── Sheet 1: แอร์ ─────────────────────────────────────────────────────────
    ws = wb['บันทึกวันล้างแอร์ 2569']
    print(f"\nImporting AC units from '{ws.title}'...")

    for row in ws.iter_rows(min_row=1, values_only=True):
        # skip header/group rows — data rows have integer in col 0 (ลำดับ)
        seq = row[0]
        if not isinstance(seq, int):
            continue

        building_code = str(row[1]).strip() if row[1] else None
        floor_name    = f"ชั้น {str(row[2]).strip()}" if row[2] else None
        dept_name     = str(row[3]).strip() if row[3] else None
        ac_type_raw   = str(row[4]).strip() if row[4] else None
        ac_code       = str(row[5]).strip() if row[5] else None
        clean_major1  = row[6]   # ล้างใหญ่ครั้งที่ 1 (date or text)

        if not all([building_code, floor_name, dept_name, ac_code]):
            continue

        ac_type   = map_type(ac_type_raw)
        ac_status = map_status(clean_major1)

        # extract capacity from type string e.g. "FCU(Wall type)/18,000BTU"
        cap = None
        if ac_type_raw and 'BTU' in ac_type_raw:
            cap = ac_type_raw.split('/')[-1].strip() if '/' in ac_type_raw else ac_type_raw
            cap = cap[:50] if cap else None

        # ── ensure building ──
        cur.execute(
            "SELECT id FROM buildings WHERE hospital_id=%s AND code=%s",
            (hospital_id, building_code)
        )
        b = cur.fetchone()
        if not b:
            cur.execute(
                "INSERT INTO buildings(hospital_id, name, code) VALUES(%s,%s,%s) RETURNING id",
                (hospital_id, f"อาคาร {building_code}", building_code)
            )
            b = cur.fetchone()
        building_id = b[0]

        # ── ensure floor ──
        cur.execute(
            "SELECT id FROM floors WHERE building_id=%s AND name=%s",
            (building_id, floor_name)
        )
        f = cur.fetchone()
        if not f:
            cur.execute(
                "INSERT INTO floors(building_id, name) VALUES(%s,%s) RETURNING id",
                (building_id, floor_name)
            )
            f = cur.fetchone()
        floor_id = f[0]

        # ── ensure department ──
        cur.execute(
            "SELECT id FROM departments WHERE floor_id=%s AND name=%s",
            (floor_id, dept_name)
        )
        d = cur.fetchone()
        if not d:
            cur.execute(
                "INSERT INTO departments(floor_id, name) VALUES(%s,%s) RETURNING id",
                (floor_id, dept_name)
            )
            d = cur.fetchone()
        dept_id = d[0]

        # ── insert ac_unit ──
        cur.execute("""
            INSERT INTO ac_units(department_id, ac_code, name, type, capacity_btu, status, pm_interval_months)
            VALUES(%s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (ac_code) DO UPDATE SET
                name=EXCLUDED.name,
                type=EXCLUDED.type,
                capacity_btu=EXCLUDED.capacity_btu,
                status=EXCLUDED.status,
                updated_at=NOW()
        """, (dept_id, ac_code, dept_name, ac_type, cap, ac_status, 2))

        total_ac += 1
        if total_ac % 100 == 0:
            print(f"  {total_ac} AC units...")

    # ── Sheet 2: พัดลม ────────────────────────────────────────────────────────
    ws2 = wb['บันทึกวันล้างพัดลม 2569']
    print(f"\nImporting Fans from '{ws2.title}'...")
    fan_seq = 1

    for row in ws2.iter_rows(min_row=1, values_only=True):
        seq = row[0]
        if not isinstance(seq, int):
            continue

        building_code = str(row[1]).strip() if row[1] else None
        floor_name    = f"ชั้น {str(row[2]).strip()}" if row[2] else None
        dept_name     = str(row[3]).strip() if row[3] else None
        fan_type_raw  = str(row[4]).strip() if row[4] else 'พัดลมดูดอากาศ'

        if not all([building_code, floor_name, dept_name]):
            continue

        # Fan ไม่มี ac_code → สร้างจาก seq
        fan_code = f"FAN-PTS1-{building_code}-{str(fan_seq).zfill(3)}"
        fan_seq += 1

        # ── ensure building/floor/dept (same logic) ──
        cur.execute(
            "SELECT id FROM buildings WHERE hospital_id=%s AND code=%s",
            (hospital_id, building_code)
        )
        b = cur.fetchone()
        if not b:
            cur.execute(
                "INSERT INTO buildings(hospital_id, name, code) VALUES(%s,%s,%s) RETURNING id",
                (hospital_id, f"อาคาร {building_code}", building_code)
            )
            b = cur.fetchone()
        building_id = b[0]

        cur.execute(
            "SELECT id FROM floors WHERE building_id=%s AND name=%s",
            (building_id, floor_name)
        )
        f = cur.fetchone()
        if not f:
            cur.execute(
                "INSERT INTO floors(building_id, name) VALUES(%s,%s) RETURNING id",
                (building_id, floor_name)
            )
            f = cur.fetchone()
        floor_id = f[0]

        cur.execute(
            "SELECT id FROM departments WHERE floor_id=%s AND name=%s",
            (floor_id, dept_name)
        )
        d = cur.fetchone()
        if not d:
            cur.execute(
                "INSERT INTO departments(floor_id, name) VALUES(%s,%s) RETURNING id",
                (floor_id, dept_name)
            )
            d = cur.fetchone()
        dept_id = d[0]

        cur.execute("""
            INSERT INTO ac_units(department_id, ac_code, name, type, status, pm_interval_months)
            VALUES(%s, %s, %s, 'Fan', 'active', 3)
            ON CONFLICT (ac_code) DO NOTHING
        """, (dept_id, fan_code, dept_name))

        total_fan += 1

    conn.commit()
    cur.close()
    conn.close()

    print(f"\nDone!")
    print(f"  AC units imported : {total_ac}")
    print(f"  Fans imported     : {total_fan}")
    print(f"  Total             : {total_ac + total_fan}")

if __name__ == '__main__':
    main()
